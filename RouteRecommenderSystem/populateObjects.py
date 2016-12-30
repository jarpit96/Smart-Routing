from routing.models import State, District, Locality
import openpyxl, googlemaps, pprint

#Create District and Locality Name and Populate With Names, Foreign Keys, Latitude and Longitude


currkey = 0
key = ['AIzaSyDmT6F29WJ9M-viNlrMzRpRPtdseHTCfoA' , 'AIzaSyAtc-2ZwV_PGZaT-TxNR1YUnicbdeCNEg0', 'AIzaSyDhic6BCNfkgzsPhsKEkZ_BvZKkzKhXEzs']
gmaps = googlemaps.Client(key=key[currkey])
pp = pprint.PrettyPrinter(indent=4)
request_cnt = 0

def createObjects():
    district_column_no = 2
    locality_column_no = 4
    state_object = {}
    global request_cnt
    wb = openpyxl.load_workbook('delhi_district_wise_localities.xlsx')
    sheet = wb.get_sheet_by_name('Sheet1')

    #Creating Delhi Object
    delhi = State.objects.create(name = 'Delhi')

    for i in range(2, sheet.max_row+1):
        district_name = sheet.cell(row=i, column=district_column_no).value
        d = District.objects.create(name = district_name, state = delhi)
        localities_in_district = str(sheet.cell(row=i, column=locality_column_no).value).split(',')
        for locality in localities_in_district:
            response = gmaps.geocode(locality, region="in")
            request_cnt+=2
            if len(response) != 0:
                Locality.objects.create(name=locality, district = d, lat = response[0]['geometry']['location']['lat'], lng = response[0]['geometry']['location']['lng'])
            else:
                continue
        state_object[district_name] = localities_in_district

    print 'District and Localities', state_object

#Update Localities with Crime Weight
def updateCrime():
    wb = openpyxl.load_workbook('crime_data_2014.xlsx')
    sheet = wb.active

    delhi_start_row_number = 815
    delhi_end_row_number = 824
    district_column_no = 2

    #Parameters For Normalization
    crime = []
    district = []

    #Calculate Total Crimes for Every District
    for i in range(delhi_start_row_number, delhi_end_row_number):
        district_name = sheet.cell(row=i, column=district_column_no).value
        district.append(district_name)
        district_crime_count = 0
        for j in range(4, sheet.max_column+1):
            district_crime_count += sheet.cell(row=i, column=j).value
        crime.append(district_crime_count)
        print district_name + ' ' +str(district_crime_count) + '\n'

    print 'Original', crime

    min_crime = min(crime)
    max_crime = max(crime)
    norm_crime = []

    #Normalize Crime Value and Update Locality Object
    for i in range (0, len(crime)):
        norm_crime.append(abs(float(crime[i]-min_crime)/(max_crime - min_crime)-1)*5)
        try:
            Locality.objects.filter(district = District.objects.get(name = district[i])).update(crime_wt = norm_crime[i])
        except District.DoesNotExist:
            continue

    print 'Normalized', norm_crime


#Add POI Weight to each Locality
def updatePOI():
    global request_cnt, currkey, gmaps
    poi_types = ['amusement_park', 'art_gallery', 'cafe', 'casino', 'hindu_temple', 'zoo', 'spa', 'restaurant',
                 'museum', 'lodging']
    states = State.objects.all()
    for state in states:
        districts = District.objects.filter(state = state)
        for district in districts:
            localities = Locality.objects.filter(district = district)
            for locality in localities:
                avg_rating = 0
                for poiType in poi_types:
                    rating_cnt = 0
                    results = gmaps.places_nearby(location=[locality.lat, locality.lng], radius=2000, type=poiType)['results']
                    request_cnt += 2
                    if request_cnt > 1000:
                        currkey = (currkey+1)%3
                        # if currkey > len(key):
                        #     print "All Keys Used"
                        #     exit()
                        gmaps = googlemaps.Client(key=key[currkey])
                        request_cnt = 0
                        print "Key updated----- current Key: ", currkey
                    for pid in results:  # For every POI Found
                        if 'rating' in pid:
                            rating_cnt += pid['rating']
                        else:
                            rating_cnt += 3.0  # assuming avg. rating = 3.0 if rating is unavailable
                    if len(results) != 0:
                        rating_cnt /= (len(results))
                    else:
                        rating_cnt = 3.0
                    avg_rating += rating_cnt
                poi_wt = avg_rating / len(poi_types)
                l = Locality.objects.get(name = locality)
                l.poi_wt = poi_wt
                l.save()
                print 'average rating for locality: ', locality, 'is: ', avg_rating / len(poi_types)


updatePOI()