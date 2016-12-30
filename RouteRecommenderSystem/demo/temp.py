import openpyxl, googlemaps, pprint, os
#from routing.models import State, District, Locality

# Read and store the Excel sheet in sheet object
wb = openpyxl.load_workbook('../delhi_district_wise_localities.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')


state_column_no = 1
district_column_no = 2
locality_column_no = 4
state_object = {}
# db_state_obj = State.objects.create(name='Delhi')
currkey = 0
key = ['AIzaSyDmT6F29WJ9M-viNlrMzRpRPtdseHTCfoA' , 'AIzaSyAtc-2ZwV_PGZaT-TxNR1YUnicbdeCNEg0', 'AIzaSyDhic6BCNfkgzsPhsKEkZ_BvZKkzKhXEzs']
gmaps = googlemaps.Client(key=key[currkey])
hasGeo = 0
noGeo = 0

request_cnt = 0

poi_types = ['amusement_park', 'art_gallery', 'cafe', 'casino', 'hindu_temple', 'zoo', 'spa', 'restaurant',
             'museum', 'lodging']
pp = pprint.PrettyPrinter(indent=4)

for i in range(2, sheet.max_row+1): #For Every District
    # district_name = sheet.cell(row=i, column=district_column_no).value
    # db_district_obj = District.objects.create(name=district_name, state=db_state_obj)
    localities_in_district = str(sheet.cell(row=i, column=locality_column_no).value).split(',')
    for locality_name in localities_in_district:  #For Every Locality
        print '################  LOCALITY  #########################: ',locality_name
        avg_rating = 0
        # Locality.objects.create(name=locality_name, district=db_district_obj, crime_wt=0.0, poi_wt=0.0,
        # traffic_wt=0.0, lat=0.0, lng=0.0)
        response = gmaps.geocode(locality_name)
        request_cnt+=2
        print "request_cnt: ", request_cnt
        if len(response) != 0:   #Coordinates Found
            loc = response[0]['geometry']['location']
            # print locality_name +" " +str(loc['lat']) + " "+str(loc['lng'])
            # print '\n'
            hasGeo = hasGeo + 1

            for poiType in poi_types: #For Every POI Type
                poi_place_ids_for_locality = []
                rating_cnt = 0
                rating_found = 0
                rating_not_found = 0
                results = gmaps.places_nearby(location=loc, radius=2000, type=poiType)['results']
                request_cnt += 2
                print "request_cnt: ", request_cnt
                if request_cnt > 10:
                    currkey += 1
                    if currkey > len(key):
                        print "All Keys Used"
                        exit()
                    gmaps = googlemaps.Client(key=key[currkey])
                    request_cnt = 0
                    print "Key updated----- current Key: ", currkey
                for pid in results: #For every POI Found
                   poi_place_ids_for_locality.append(pid['place_id'])
                   if 'rating' in pid:
                       rating_cnt += pid['rating']
                       rating_found +=1
                   else:
                       rating_cnt += 2.5 #assuming avg. rating = 2.5 if rating is unavailable
                       rating_not_found += 1

                if len(results) != 0:
                    rating_cnt = rating_cnt/(len(results))
                else:
                    rating_cnt = 2.5
                avg_rating += rating_cnt
                print 'Avg rating for type: ', poiType, 'is: ' + str(rating_cnt)+'\n'
                print 'Found', rating_found, 'Not FOund', rating_not_found
            # for place_id in poi_place_ids_for_locality:
            #     rating_cnt += gmaps.place(place_id=place_id)
        else:
            noGeo = noGeo + 1
            print 'not found', locality_name + '\n'

        print 'average rating for locality: ', locality_name, 'is: ', avg_rating/10
            # state_object[district_name] = localities_in_district

# print state_object
# print hasGeo
# print '\n'
# print noGeo
#
