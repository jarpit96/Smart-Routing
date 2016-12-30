import openpyxl

wb = openpyxl.load_workbook('crime_data_2014.xlsx')
sheet = wb.active

delhi_start_row_number = 815
delhi_end_row_number = 824
district_column_no = 2

crime = []
district = []
s = 0 #Standard Deviation
m = 0 #Mean

for i in range(delhi_start_row_number, delhi_end_row_number):
    district_name = sheet.cell(row=i, column=district_column_no).value
    district.append(district_name)
    district_crime_count = 0
    for j in range(4, sheet.max_column+1):
        district_crime_count += sheet.cell(row=i, column=j).value
    crime.append(district_crime_count)
    m += district_crime_count
    print district_name + ' ' +str(district_crime_count) + '\n'
m = m/len(crime)

print 'Original', crime

for c in crime:
    s += pow(c-m, 2)
s = s/len(crime)
s = pow(s, 0.5)

for i in range(0, len(crime)):
    crime[i] = (float)(crime[i]-m)/s

print 'Normalized', crime