import openpyxl
from routing.models import State, District, Locality

# Read and store the Excel sheet in sheet object
wb = openpyxl.load_workbook('delhi_district_wise_localities.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

state_column_no = 1
district_column_no = 2
locality_column_no = 4
state_object = {}
db_state_obj = State.objects.create(name='Delhi')

for i in range(2, sheet.max_row):
    district_name = sheet.cell(row=i, column=district_column_no).value
    db_district_obj = District.objects.create(name=district_name, state=db_state_obj)
    localities_in_district = str(sheet.cell(row=i, column=locality_column_no).value).split(',')
    for locality_name in localities_in_district:
        Locality.objects.create(name=locality_name, district=db_district_obj, crime_wt=0, poi_wt=0,
                                traffic_wt=0, lat=0, lng=0)
    state_object[district_name] = localities_in_district

print state_object
