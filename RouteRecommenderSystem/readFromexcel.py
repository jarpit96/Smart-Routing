import openpyxl

# Read and store the Excel sheet in sheet object
wb = openpyxl.load_workbook('delhi_district_wise_localities.xlsx')
sheet = wb.get_sheet_by_name('Sheet1')

state_column_no = 1
district_column_no = 2
locality_column_no = 4
state_object = {}

for i in range(2, sheet.max_row):
    district_name = sheet.cell(row=i, column=district_column_no).value
    localities_in_district = str(sheet.cell(row=i, column=locality_column_no).value).split(',')
    state_object[district_name] = localities_in_district

print state_object
